"""数据处理模块（优化版）"""
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import yaml
from pathlib import Path
import tqdm
from src.common.push_data2kafka import push_data2kafka
from src.common.db import MySQLConnector
from src.common.logger import get_logger

# 读取配置
CONFIG_PATH = Path(__file__).resolve().parents[2] / "config.yaml"

# 检查配置文件是否存在
if not CONFIG_PATH.exists():
    # 如果标准路径不存在，则尝试相对路径查找
    CONFIG_PATH = Path("config.yaml").resolve()
    if not CONFIG_PATH.exists():
        # 如果还是找不到，使用默认配置
        DATA_PATH = Path("../data/prod_data.xlsx").resolve()  # 相对于src/data_processing目录的路径
        DIRECT_EVALUATION_PATH = Path("../data/voc线上数据.xlsx").resolve()  # 直接评估文件路径
        DATA_PROCESSING_MODE = "direct"  # 默认使用直接模式
    else:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f)
        # 使用项目根目录作为基准来构建数据路径
        PROJECT_ROOT = Path(__file__).resolve().parents[2]  # 回到项目根目录
        DATA_PATH = PROJECT_ROOT / config["data"]["prod_data_path"]
        DIRECT_EVALUATION_PATH = PROJECT_ROOT / config["data"]["direct_evaluation_path"]
        DATA_PROCESSING_MODE = config["evaluation"].get("data_processing_mode", "direct")
else:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    # 使用项目根目录作为基准来构建数据路径
    PROJECT_ROOT = Path(__file__).resolve().parents[2]  # 回到项目根目录
    DATA_PATH = PROJECT_ROOT / config["data"]["prod_data_path"]
    DIRECT_EVALUATION_PATH = PROJECT_ROOT / config["data"]["direct_evaluation_path"]
    DATA_PROCESSING_MODE = config["evaluation"].get("data_processing_mode", "direct")

# 初始化日志
logger = get_logger(__name__)

def read_excel(filepath: Path):
    """
    读取Excel文件（增强异常处理+日志+多工作表支持）
    :param filepath: 数据文件路径（Path对象）
    :return: 数据列表（字典格式）
    """
    if not filepath.exists():
        logger.error(f"Excel文件不存在: {filepath}")
        raise FileNotFoundError(f"File not found: {filepath}")
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active  # 默认使用活动工作表
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        data = []
        for row in rows[1:]:
            data_dict = dict(zip(header, row))
            data.append(data_dict)
        return data
    except InvalidFileException as e:
        logger.error(f"Excel文件格式错误: {str(e)}", exc_info=True)
        raise
    except Exception as e:
        logger.error(f"读取Excel失败: {str(e)}", exc_info=True)
        raise

def build_labels_from_direct_data(filepath: Path):
    """
    从直接包含y_true和y_pred的Excel文件中构建标签数据
    :param filepath: 包含y_true和y_pred列的Excel文件路径
    :return: y_true, y_pred 列表
    """
    try:
        data = read_excel(filepath)
        
        y_true = []
        y_pred = []
        for item in data:
            true_value = item["voc_tendencies_manually"]
            pred_value = item["voc_tendencies"]
            
            # 跳过None值
            if true_value is None or pred_value is None:
                logger.warning(f"发现空标签: y_true={true_value}, y_pred={pred_value}, 在数据项: {item}")
                continue
                
            # 标准化标签值
            true_label = str(true_value).lower().strip()
            pred_label = str(pred_value).lower().strip()
            
            # 处理中文标签
            if true_label in ['负面', '消极', '负向', 'neg', 'negative', 'negative ']:
                true_label = "negative"
            elif true_label in ['正面', '积极', '正向', 'pos', 'positive', 'positive ']:
                true_label = "positive"
            elif true_label in ['中性', '中立', 'neu', 'neutral', 'neutral ']:
                true_label = "neutral"
            else:
                logger.warning(f"未知的真实标签值: {true_label}, 将其转换为neutral")
                true_label = "neutral"
            
            if pred_label in ['负面', '消极', '负向', 'neg', 'negative', 'negative ']:
                pred_label = "negative"
            elif pred_label in ['正面', '积极', '正向', 'pos', 'positive', 'positive ']:
                pred_label = "positive"
            elif pred_label in ['中性', '中立', 'neu', 'neutral', 'neutral ']:
                pred_label = "neutral"
            else:
                logger.warning(f"未知的预测标签值: {pred_label}, 将其转换为neutral")
                pred_label = "neutral"
                
            y_true.append(true_label)
            y_pred.append(pred_label)
        
        logger.info(f"从直接数据构建标签完成 | y_true条数: {len(y_true)}, y_pred条数: {len(y_pred)}")
        return y_true, y_pred
        
    except Exception as e:
        logger.error("从直接数据构建标签失败", exc_info=True)
        raise

def build_real_label_data(data_path: Path):
    """构建真实标签数据"""
    try:
        data = read_excel(data_path)
        real_label_data = []
        for item in data:
            # 校验标签字段是否存在
            if "voc_tendencies_manually" not in item:
                logger.error("Excel缺少字段: voc_tendencies_manually")
                raise KeyError("Missing field: voc_tendencies_manually")
            label = item["voc_tendencies_manually"]
            if label not in ["negative", "positive", "neutral"]:
                logger.warning(f"无效标签: {label} | 数据ID: {item.get('unique_id', '未知')}")
            real_label_data.append(label)
        
        logger.info(f"构建真实标签完成 | 条数: {len(real_label_data)}")
        return real_label_data
    except Exception as e:
        logger.error("构建真实标签失败", exc_info=True)
        raise

def build_pre_label_data(origin_data):
    """构建预测标签数据"""
    if not origin_data:
        logger.error("原始数据为空，无法构建预测标签")
        raise ValueError("origin_data is empty")
    
    pre_label_data = []
    try:
        with MySQLConnector() as query_sql:
            for item in tqdm.tqdm(origin_data, desc="构建预测标签"):
                # 校验必要字段
                if "unique_id" not in item:
                    logger.warning("数据缺少unique_id，跳过推送")
                    continue
                # 推送数据到Kafka
                push_data2kafka(item)
                # 查询预测结果
                result = query_sql.execute_query(item=item)
                if result:
                    pre_label = result[0]['voc_tendencies']
                    pre_label_data.append(pre_label)
                else:
                    pre_label_data.append("neutral")  # 无结果时默认中性
        logger.info(f"构建预测标签完成 | 条数: {len(pre_label_data)}")
        return pre_label_data
    except Exception as e:
        logger.error("构建预测标签失败", exc_info=True)
        raise

def build_labels(data_path: Path, processing_mode: str = None):
    """
    根据指定模式构建标签数据
    :param data_path: 数据文件路径
    :param processing_mode: 处理模式 ("direct" 或 "sql_query")
    :return: y_true, y_pred 元组
    """
    mode = processing_mode or DATA_PROCESSING_MODE
    
    if mode == "direct":
        logger.info("使用直接数据模式构建标签")
        return build_labels_from_direct_data(DIRECT_EVALUATION_PATH)
    elif mode == "sql_query":
        logger.info("使用SQL查询模式构建标签")
        origin_data = read_excel(data_path)
        y_true = build_real_label_data(data_path)
        y_pred = build_pre_label_data(origin_data)
        return y_true, y_pred
    else:
        raise ValueError(f"不支持的数据处理模式: {mode}")

if __name__ == "__main__":
    try:
        # 根据配置决定使用哪种模式
        y_true, y_pred = build_labels(DATA_PATH)
        logger.info(f"测试 | 真实标签条数: {len(y_true)} | 预测标签条数: {len(y_pred)}")
    except Exception as e:
        logger.error("测试数据处理失败", exc_info=True)