from openpyxl import load_workbook
import os



data_path = os.getcwd() + "\\voc线上数据.xlsx"
# print(f"data_path: {data_path}")

wb = load_workbook(data_path, read_only=True, data_only=True)
ws = wb.active  # 默认使用活动工作表

data = []
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
for row in rows[1:4]:
    data_dict = dict(zip(header, row))
    data.append(data_dict)


print(data)